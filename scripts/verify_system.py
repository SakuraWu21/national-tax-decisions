#!/usr/bin/env python3
"""对交付物执行结构、去重、关联、数量和格式完整性检查。"""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

from jsonschema import Draft202012Validator
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "全国税务处理及行政处罚决定书汇总.xlsx"
CSV = ROOT / "data" / "全国税务处理及行政处罚决定书汇总.csv"
SHEETS = ["文书汇总", "完整文书", "公告送达及文号线索", "待核验记录", "失效链接", "每日运行日志"]
PUBLIC_JSON = ROOT / "public" / "data" / "tax-decisions.json"
PUBLIC_STATUS = ROOT / "public" / "data" / "update-status.json"
PUBLIC_XLSX = ROOT / "public" / "downloads" / XLSX.name
PUBLIC_CSV = ROOT / "public" / "downloads" / CSV.name
SCHEMA = ROOT / "config" / "tax-decisions.schema.json"


def fail(message: str) -> None:
    raise AssertionError(message)


def main() -> int:
    if not XLSX.exists() or not CSV.exists():
        fail("Excel 或 CSV 不存在")
    workbook = load_workbook(XLSX, read_only=False, data_only=False)
    if workbook.sheetnames != SHEETS:
        fail(f"工作表不正确：{workbook.sheetnames}")
    summary = workbook["文书汇总"]
    headers = [cell.value for cell in summary[1]]
    records = []
    for row in summary.iter_rows(min_row=2):
        if not any(cell.value not in (None, "") for cell in row):
            continue
        item = {}
        for index, header in enumerate(headers):
            cell = row[index]
            value = cell.hyperlink.target if cell.hyperlink else cell.value
            item[header] = value if value is not None else ""
        records.append(item)

    with CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    if len(records) != len(csv_rows):
        fail(f"Excel/CSV 数量不一致：{len(records)} != {len(csv_rows)}")
    for path in (PUBLIC_JSON, PUBLIC_STATUS, PUBLIC_XLSX, PUBLIC_CSV, SCHEMA):
        if not path.exists():
            fail(f"前端交付文件不存在：{path}")
    public_rows = json.loads(PUBLIC_JSON.read_text(encoding="utf-8"))
    schema = json.loads(SCHEMA.read_text(encoding="utf-8"))
    schema_errors = sorted(Draft202012Validator(schema).iter_errors(public_rows), key=lambda item: list(item.path))
    if schema_errors:
        fail(f"JSON Schema 校验失败：{schema_errors[0].message}")
    if len(public_rows) != len(records):
        fail(f"JSON/Excel 数量不一致：{len(public_rows)} != {len(records)}")
    public_ids = [str(record["id"]) for record in public_rows]
    if len(public_ids) != len(set(public_ids)):
        fail("前端 JSON 存在重复文书唯一ID")
    status = json.loads(PUBLIC_STATUS.read_text(encoding="utf-8"))
    if status.get("total") != len(public_rows):
        fail("update-status.json 总数与记录数量不一致")
    if PUBLIC_XLSX.read_bytes() != XLSX.read_bytes() or PUBLIC_CSV.read_bytes() != CSV.read_bytes():
        fail("公开下载文件与数据源文件不一致")

    ids = [str(record["文书唯一ID"]) for record in records]
    if len(ids) != len(set(ids)):
        fail("存在重复文书唯一ID")
    type_numbers = [
        (str(record["文书类型"]), str(record["决定书文号"]).replace(" ", ""))
        for record in records if record["决定书文号"]
    ]
    if len(type_numbers) != len(set(type_numbers)):
        fail("存在重复的文书类型+决定书文号")

    grouped: dict[str, list[dict]] = {}
    for record in records:
        grouped.setdefault(str(record["案件组ID"]), []).append(record)
    paired_groups = 0
    for group in grouped.values():
        types = {record["文书类型"] for record in group}
        if {"税务处理决定书", "税务行政处罚决定书"} <= types:
            paired_groups += 1
            if len({record["案件组ID"] for record in group}) != 1:
                fail("关联文书案件组ID不一致")
            if not all(record["关联处理决定书文号"] and record["关联处罚决定书文号"] for record in group):
                fail("关联文号未双向写入")

    if summary.freeze_panes != "A2":
        fail("文书汇总未冻结首行")
    if not summary.auto_filter.ref:
        fail("文书汇总未开启筛选")
    for sheet_name in SHEETS:
        sheet = workbook[sheet_name]
        if sheet.max_row < 1 or sheet["A1"].value in (None, ""):
            fail(f"{sheet_name} 缺少表头")
    log_sheet = workbook["每日运行日志"]
    if log_sheet.max_row < 2:
        fail("每日运行日志为空")

    last_log = {log_sheet.cell(1, col).value: log_sheet.cell(log_sheet.max_row, col).value for col in range(1, log_sheet.max_column + 1)}
    report = {
        "xlsx_opened": True,
        "excel_records": len(records),
        "csv_records": len(csv_rows),
        "unique_ids": len(set(ids)),
        "json_records": len(public_rows),
        "json_schema_valid": True,
        "download_files_match": True,
        "paired_case_groups": paired_groups,
        "pending_records": sum(1 for record in records if record["核验状态"] == "待核验"),
        "invalid_link_records": sum(1 for record in records if record["页面当前状态"] not in ("正常", "附件正常")),
        "log_rows": log_sheet.max_row - 1,
        "last_run_success": last_log.get("运行是否成功"),
        "all_required_sheets": True,
        "freeze_and_filter": True,
    }
    workbook.close()
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print(f"VERIFICATION FAILED: {type(exc).__name__}: {exc}", file=sys.stderr)
        sys.exit(1)
